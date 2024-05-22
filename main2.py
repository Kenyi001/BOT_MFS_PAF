import os
import sys
from typing import Tuple, Optional, Callable
from PyQt5.QtCore import Qt, QThread, pyqtSignal
from PyQt5.QtGui import QIcon, QPixmap
from PyQt5.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QPushButton, QFileDialog,
    QMessageBox, QInputDialog, QLineEdit, QHBoxLayout, QFormLayout,
    QDialog, QLabel, QTextEdit, QProgressDialog
)
from openpyxl.reader.excel import load_workbook

from gui import ConfigurationDialog
from selenium_scripts.new_PAF_PTM import ejecutar_automatizacion_new
from selenium_scripts.extraccion_datos_web import extraer_datos_web
from config import URL_INICIO_SESION

class Worker(QThread):
    finished = pyqtSignal()
    progress = pyqtSignal(int)

    def __init__(self, func: Callable, *args) -> None:
        super().__init__()
        self.func = func
        self.args = args

    def run(self) -> None:
        self.func(*self.args)
        self.finished.emit()

class SelectableInfoDialog(QDialog):
    def __init__(self, parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        self.setWindowTitle('Información')
        self.setMinimumSize(400, 300)
        self.layout = QVBoxLayout(self)
        self.info_text_edit = QTextEdit(self)
        self.info_text_edit.setReadOnly(True)
        self.info_text_edit.setStyleSheet("font-size: 16px;")
        self.layout.addWidget(self.info_text_edit)
        self.ok_button = QPushButton('OK', self)
        self.ok_button.clicked.connect(self.accept)
        self.layout.addWidget(self.ok_button)

    def set_info_text(self, text: str) -> None:
        self.info_text_edit.setHtml(text)

class CredentialsDialog(QDialog):
    def __init__(self, parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        self.setWindowTitle('Solicitar credenciales')
        self.layout = QFormLayout(self)
        self.username_line_edit = QLineEdit(self)
        self.username_line_edit.setMinimumWidth(200)
        self.password_line_edit = QLineEdit(self)
        self.password_line_edit.setEchoMode(QLineEdit.Password)
        self.show_password_button = QPushButton('Mostrar contraseña', self)
        self.show_password_button.setCheckable(True)
        self.show_password_button.clicked.connect(self.toggle_password_visibility)
        self.layout.addRow(QLabel("Nombre de usuario:"), self.username_line_edit)
        self.layout.addRow(QLabel("Contraseña:"), self.password_line_edit)
        self.layout.addRow(self.show_password_button)
        self.ok_button = QPushButton('OK', self)
        self.ok_button.clicked.connect(self.accept)
        self.layout.addRow(self.ok_button)

    def toggle_password_visibility(self, checked: bool) -> None:
        self.password_line_edit.setEchoMode(QLineEdit.Normal if checked else QLineEdit.Password)

    def get_credentials(self) -> Tuple[str, str]:
        return self.username_line_edit.text(), self.password_line_edit.text()

class App(QWidget):
    def __init__(self) -> None:
        super().__init__()
        self.setWindowTitle('BOT PTM - ASFI')
        self.resize(350, 110)
        self.init_ui()

    def init_ui(self) -> None:
        layout = QVBoxLayout()

        # Crear un layout horizontal para el logo y el botón de información
        logo_info_layout = QHBoxLayout()
        logo = QLabel(self)
        logo.setPixmap(QPixmap('D:\Comercial_-_Daniel\BOT_MFS_PAF\Static\logo_home.png'))
        logo_info_layout.addWidget(logo)

        self.btn_info = QPushButton('!', self)
        self.btn_info.clicked.connect(self.show_general_info)
        self.btn_info.setFixedSize(30, 30)
        logo_info_layout.addWidget(self.btn_info)

        layout.addLayout(logo_info_layout)

        self.btn_request_paths = QPushButton('Solicitar rutas')
        self.btn_request_paths.clicked.connect(self.open_configuration_dialog)
        layout.addWidget(self.btn_request_paths)

        layout.addLayout(self.create_paf_ptm_buttons())
        layout.addLayout(self.create_web_data_buttons())
        self.setLayout(layout)

    def create_paf_ptm_buttons(self) -> QHBoxLayout:
        layout = QHBoxLayout()
        self.btn_new_paf_ptm = QPushButton('New PAF/PTM')
        self.btn_new_paf_ptm.clicked.connect(self.open_file_dialog_new_paf_ptm)
        layout.addWidget(self.btn_new_paf_ptm)
        self.btn_info_new_paf_ptm = QPushButton('Inf.')
        self.btn_info_new_paf_ptm.clicked.connect(self.show_data_info_new_paf_ptm)
        self.btn_info_new_paf_ptm.setFixedSize(50, 30)
        layout.addWidget(self.btn_info_new_paf_ptm)
        return layout

    def show_general_info(self) -> None:
        dialog = SelectableInfoDialog(self)
        dialog.set_info_text(
            "<h1>Descripción de la Aplicación</h1>"
            "<p>Esta aplicación está diseñada para automatizar tareas específicas "
            "relacionadas con la gestión de PTM (Punto de Transferencia Monetaria) "
            "y la extracción de datos web. La aplicación ofrece una interfaz gráfica "
            "amigable que permite al usuario interactuar fácilmente con sus funcionalidades.</p>"
            "<h2>Funcionalidades Principales</h2>"
            "<ol>"
            "<li><b>Solicitar Rutas:</b> Permite al usuario configurar las rutas necesarias "
            "para el funcionamiento del bot.</li>"
            "<li><b>Automatización de Nuevo PAF/PTM:</b> Realiza la creación de nuevos registros "
            "de PTM basándose en un archivo Excel proporcionado por el usuario. Cada registro "
            "toma aproximadamente 56 segundos en procesarse.</li>"
            "<li><b>Extracción de Datos Web:</b> Extrae datos específicos desde una web utilizando "
            "la información proporcionada en un archivo Excel. Cada extracción toma aproximadamente "
            "30 segundos en completarse.</li>"
            "</ol>"
            "<h2>Ventanas de Información y Confirmación</h2>"
            "<ul>"
            "<li>La aplicación puede mostrar ventanas emergentes con información relevante y detallada "
            "sobre los procesos y sus requisitos.</li>"
            "<li>Al finalizar una tarea, se muestra una ventana de confirmación indicando que el proceso "
            "se ha completado exitosamente.</li>"
            "<li>En caso de errores, se muestra una ventana de alerta notificando al usuario sobre los "
            "problemas ocurridos durante la ejecución.</li>"
            "</ul>"
            "<h2>Interacción del Usuario</h2>"
            "<ul>"
            "<li>El usuario puede cargar archivos Excel necesarios para la ejecución de las tareas.</li>"
            "<li>Se solicita al usuario ingresar sus credenciales para acceder a las funciones de automatización.</li>"
            "</ul>"
            "<h2>Ejecución de la Aplicación</h2>"
            "<p>La aplicación se ejecuta localmente en la PC del usuario, proporcionando una herramienta "
            "robusta y eficiente para gestionar las tareas de automatización de PTM y extracción de datos.</p>"
            "<br>"
            "<h2>Repositorio por: https://github.com/Kenyi001/BOT_MFS_PAF/tree/Desktop-App</h2>"
            
        )
        dialog.exec_()

    def show_data_info_new_paf_ptm(self) -> None:
        dialog = SelectableInfoDialog(self)
        dialog.set_info_text(
            "<h1>Información para Nuevo PAF/PTM</h1>"
            "<p>Para introducir nuevos PTM, asegúrate de que el archivo Excel (.xlsx) "
            "contenga las siguientes columnas:</p>"
            "<ul>"
            "<li>mef</li>"
            "<li>nombre_corresponsable</li>"
            "<li>nro_carnet</li>"
            "<li>nombre_del_negocio</li>"
            "<li>direccion</li>"
            "<li>camara_de_seguridad</li>"
            "<li>linea_personal_corresponsable</li>"
            "<li>horario_lunes</li>"
            "<li>horario_martes</li>"
            "<li>horario_miercoles</li>"
            "<li>horario_jueves</li>"
            "<li>horario_viernes</li>"
            "<li>horario_sabado</li>"
            "<li>horario_domingo</li>"
            "<li>localidad</li>"
            "<li>departamento</li>"
            "<li>lat_gps</li>"
            "<li>lng_gps</li>"
            "</ul>"
        )
        dialog.exec_()

    def show_data_info_web_data_extraction(self) -> None:
        dialog = SelectableInfoDialog(self)
        dialog.set_info_text(
            "<h1>Información para Extracción de Datos Web</h1>"
            "<p>Para 'Extracción de Datos Web', se necesita la siguiente columna en el archivo Excel:</p>"
            "<ul>"
            "<li>MEF</li>"
            "</ul>"
        )
        dialog.exec_()

    def create_web_data_buttons(self) -> QHBoxLayout:
        layout = QHBoxLayout()
        self.btn_extraccion_datos_web = QPushButton('Extracción de Datos Web')
        self.btn_extraccion_datos_web.clicked.connect(self.open_file_dialog_web_data_extraction)
        layout.addWidget(self.btn_extraccion_datos_web)
        self.btn_info_extraccion_datos_web = QPushButton('Inf.')
        self.btn_info_extraccion_datos_web.clicked.connect(self.show_data_info_web_data_extraction)
        self.btn_info_extraccion_datos_web.setFixedSize(50, 30)
        layout.addWidget(self.btn_info_extraccion_datos_web)
        return layout

    def open_configuration_dialog(self) -> None:
        config_dialog = ConfigurationDialog(self)
        config_dialog.exec_()

    def request_credentials(self) -> Tuple[Optional[str], Optional[str]]:
        dialog = CredentialsDialog(self)
        result = dialog.exec_()
        if result == QDialog.Accepted:
            return dialog.get_credentials()
        return None, None

    def open_file_dialog_new_paf_ptm(self) -> None:
        file_path, _ = QFileDialog.getOpenFileName(self, 'Seleccionar archivo', '', 'Excel Files (*.xlsx *.xls)')
        if file_path:
            username, password = self.request_credentials()
            if username and password:
                workbook = load_workbook(filename=file_path, read_only=True)
                sheet_names = workbook.sheetnames
                sheet_name, ok = QInputDialog.getItem(
                    self, 'Solicitar nombre de hoja',
                    'Por favor, selecciona el nombre de la hoja de cálculo:',
                    sheet_names, 0, False
                )
                if ok and sheet_name:
                    self.run_task_with_progress(
                        ejecutar_automatizacion_new,
                        URL_INICIO_SESION, file_path, sheet_name, username, password,
                        label="Ejecutando 'New PAF/PTM'..."
                    )
                    self.show_confirmation_message("La automatización 'New PAF/PTM' se completó con éxito.")

    def open_file_dialog_web_data_extraction(self) -> None:
        file_path, _ = QFileDialog.getOpenFileName(self, 'Seleccionar archivo', '', 'Excel Files (*.xlsx *.xls)')
        if file_path:
            username, password = self.request_credentials()
            if username and password:
                self.run_task_with_progress(
                    extraer_datos_web,
                    URL_INICIO_SESION, file_path, username, password,
                    label="Ejecutando extracción de datos web..."
                )
                self.show_confirmation_message("La extracción de datos web se completó con éxito.")

    def run_task_with_progress(self, func: Callable, *args, label: str) -> None:
        progress = QProgressDialog(label, "Cancelar", 0, 0, self)
        progress.setWindowModality(Qt.WindowModal)

        worker = Worker(func, *args)
        worker.finished.connect(progress.close)
        worker.finished.connect(worker.deleteLater)
        worker.start()

        progress.exec_()

    def show_confirmation_message(self, message: str) -> None:
        msg_box = QMessageBox()
        msg_box.setIcon(QMessageBox.Information)
        msg_box.setText(message)
        msg_box.setWindowTitle("Confirmación")
        msg_box.setStandardButtons(QMessageBox.Ok)
        msg_box.exec_()

if __name__ == '__main__':
    app = QApplication(sys.argv)

    exe_path = sys.executable
    dir_path = os.path.dirname(exe_path)
    icon_path = os.path.join(dir_path, 'Tigo_Money_BOT_icon.ico')

    if os.path.exists(icon_path):
        app.setWindowIcon(QIcon(icon_path))
    else:
        print(f"El archivo de icono no se encontró en la ruta: {icon_path}")

    app.setStyleSheet("""
        QWidget {
            background-color: #333;
            color: #fff;
            border-radius: 5px;
        }
        QPushButton {
            background-color: #555;
            border: none;
            border-radius: 5px;
            padding: 5px 10px;
        }
        QPushButton:hover {
            background-color: #777;
        }
        QPushButton:pressed {
            background-color: #999;
        }
        QLineEdit {
            background-color: #555;
            border: none;
            border-radius: 5px;
            padding: 5px 10px;
        }
    """)
    window = App()
    window.show()
    sys.exit(app.exec_())

# ejecutar comando en terminal para el .exe:
# pyinstaller --onefile --windowed --icon=D:\Comercial_-_Daniel\BOT_MFS_PAF\Static\Tigo_Money_BOT_ico.ico --name=Bot_PTM_Asfi main2.py

